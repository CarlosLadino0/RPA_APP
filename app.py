from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QLineEdit, QPushButton, QMessageBox,
                             QFileDialog, QVBoxLayout, QComboBox, QCheckBox, QScrollArea, QMainWindow)
from PyQt5.QtGui import QMovie, QIntValidator, QRegExpValidator, QPixmap
from openpyxl.styles import Font, Alignment, Border, Side
from PyQt5.QtCore import Qt, QRegExp, QTimer
from openpyxl import Workbook, load_workbook
from procesador import procesar_documentos
import subprocess
import json
import sys
import os
import re

class FormularioApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('COTIZADOR AUTOMÁTICO A.G')
        self.setGeometry(50, 50, 765, 600)
        print("App abierta correctamente")

        self.cargar_estilos()
        layout_principal = QVBoxLayout()

        #Logo
        self.logo = QLabel(self)
        pixmap = QPixmap("recursos/LOGO.png")
        self.logo.setPixmap(pixmap)
        self.logo.setAlignment(Qt.AlignCenter)
        layout_principal.addWidget(self.logo)

        #Cédula
        self.label_cedula = QLabel('Adjuntar Cédula (PDF O IMAGEN):')
        self.boton_cedula = QPushButton('Seleccionar archivo')
        self.boton_cedula.clicked.connect(self.seleccionar_archivo_cedula)
        self.ruta_cedula = QLineEdit()
        self.ruta_cedula.setPlaceholderText('No se ha seleccionado ningún archivo')
        self.ruta_cedula.setReadOnly(True)
        layout_cedula = QVBoxLayout()
        layout_cedula.addWidget(self.label_cedula)
        layout_cedula.addWidget(self.boton_cedula)
        layout_cedula.addWidget(self.ruta_cedula)

        #Tarjeta de propiedad
        self.label_tarjeta = QLabel('Adjuntar tarjeta de propiedad (PDF O IMAGEN):')
        self.boton_tarjeta = QPushButton('Seleccionar archivo')
        self.boton_tarjeta.clicked.connect(self.seleccionar_archivo_tarjeta)
        self.ruta_tarjeta = QLineEdit()
        self.ruta_tarjeta.setPlaceholderText('No se ha seleccionado ningún archivo')
        self.ruta_tarjeta.setReadOnly(True)
        layout_tarjeta = QVBoxLayout()
        layout_tarjeta.addWidget(self.label_tarjeta)
        layout_tarjeta.addWidget(self.boton_tarjeta)
        layout_tarjeta.addWidget(self.ruta_tarjeta)

        #Vehículo nuevo
        self.checkbox_nuevo = QCheckBox("¿Vehículo nuevo (0KM)?")
        self.checkbox_nuevo.stateChanged.connect(self.toggle_nuevo)
        self.boton_factura = QPushButton('Subir factura (PDF)')
        self.boton_factura.clicked.connect(self.seleccionar_factura)
        self.boton_factura.hide()
        self.ruta_factura = QLineEdit()
        self.ruta_factura.setPlaceholderText('No se ha seleccionado ninguna factura')
        self.ruta_factura.setReadOnly(True)
        self.ruta_factura.hide()
        layout_nuevo = QVBoxLayout()
        layout_nuevo.addWidget(self.checkbox_nuevo)
        layout_nuevo.addWidget(self.boton_factura)
        layout_nuevo.addWidget(self.ruta_factura)

        #Zona de circulación
        self.label_departamento = QLabel('Zona de circulación:')
        self.combo_departamento = QComboBox()
        self.combo_departamento.addItem('Seleccione un departamento')
        self.combo_departamento.currentIndexChanged.connect(self.actualizar_ciudades)
        self.label_ciudad = QLabel('Ciudad:')
        self.combo_ciudad = QComboBox()
        self.combo_ciudad.addItem('Seleccione una ciudad')
        layout_zona = QVBoxLayout()
        layout_zona.addWidget(self.label_departamento)
        layout_zona.addWidget(self.combo_departamento)
        layout_zona.addWidget(self.label_ciudad)
        layout_zona.addWidget(self.combo_ciudad)

        #Género del propietario
        self.label_genero = QLabel('Genero del propietario:')
        self.combo_genero = QComboBox()
        self.combo_genero.addItem('Seleccione un genero')
        self.combo_genero.addItem('MASCULINO')
        self.combo_genero.addItem('FEMENINO')
        layout_genero = QVBoxLayout()
        layout_genero.addWidget(self.label_genero)
        layout_genero.addWidget(self.combo_genero)

        #Oneroso
        self.checkbox_oneroso = QCheckBox("¿Beneficiario Oneroso?", self)
        self.checkbox_oneroso.stateChanged.connect(self.toggle_oneroso)
        self.combo_oneroso = QComboBox()
        self.combo_oneroso.addItem('Seleccione una opción')
        self.combo_oneroso.hide()
        layout_oneroso = QVBoxLayout()
        layout_oneroso.addWidget(self.checkbox_oneroso)
        layout_oneroso.addWidget(self.combo_oneroso)
        self.cargar_oneroso()

        #Correo electrónico
        self.label_correo = QLabel('Correo electrónico:')
        self.campo_correo = QLineEdit()
        self.campo_correo.setPlaceholderText('example@example.com')
        layout_correo = QVBoxLayout()
        layout_correo.addWidget(self.label_correo)
        layout_correo.addWidget(self.campo_correo)
        self.campo_correo.textChanged.connect(self.validar_correo)

        # Boton procesar
        self.boton_procesar = QPushButton('Procesar')
        self.boton_procesar.clicked.connect(self.procesar_formulario)

        self.label_estado = QLabel('', self)
        self.label_estado.setAlignment(Qt.AlignCenter)
        self.label_estado.setStyleSheet("color: blue; font-weight: bold;")
        self.label_estado.hide()

        layout_principal.addLayout(layout_cedula)
        layout_principal.addLayout(layout_tarjeta)
        layout_principal.addLayout(layout_zona)
        layout_principal.addLayout(layout_genero)
        layout_principal.addLayout(layout_correo)
        layout_principal.addLayout(layout_nuevo)
        layout_principal.addLayout(layout_oneroso)
        layout_principal.addWidget(self.boton_procesar)
        layout_principal.addWidget(self.label_estado)

        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        widget_contenido = QWidget()
        widget_contenido.setLayout(layout_principal)
        scroll_area.setWidget(widget_contenido)
        layout_ventana = QVBoxLayout(self)
        layout_ventana.addWidget(scroll_area)

        self.setLayout(layout_ventana)

        self.departamentos = self.cargar_departamentos()

        for departamento in self.departamentos:
            self.combo_departamento.addItem(departamento['departamento'])

    def wheelEvent(self, event):
        widget = QApplication.focusWidget()
        if isinstance(widget, QComboBox):
            if widget.hasFocus():
                super().wheelEvent(event)  
            else:
                event.ignore()  
        else:
            super().wheelEvent(event)

    def cargar_estilos(self):
        ruta_estilos = os.path.join(os.path.dirname(__file__), 'estilos', 'styles.css')

        with open(ruta_estilos, 'r') as file:
            self.setStyleSheet(file.read())

    def cargar_departamentos(self):
        try:
            with open('recursos/ciudades.json', 'r', encoding='utf-8') as archivo:
                return json.load(archivo)
        except FileNotFoundError:
            self.mostrar_alerta('Error', 'No se encontró el archivo de ciudades.', "Critical")
            return []
        except json.JSONDecodeError:
            self.mostrar_alerta('Error', 'Error al leer el archivo de ciudades.', "Critical")
            return []

    def cargar_oneroso(self):
        try:
            with open('recursos/oneroso.json', 'r', encoding='utf-8') as archivo:
                data = json.load(archivo)
                return data[0].get('bancos', [])
        except FileNotFoundError:
            self.mostrar_alerta('Error', 'No se encontró el archivo de onerosos')
            return []
        except json.JSONDecodeError:
            self.mostrar_alerta('Error', 'Error al leer el archivo de onerosos')
            return []

    def cargar_opciones_oneroso(self):
        bancos = self.cargar_oneroso()  
        self.combo_oneroso.clear()
        self.combo_oneroso.addItem('Seleccione una opción')  
        for banco in bancos:
            self.combo_oneroso.addItem(banco)

    def toggle_nuevo(self, state):
        if state == Qt.Checked:
            self.boton_factura.show()
            self.ruta_factura.show()
        else:
            self.boton_factura.hide()
            self.ruta_factura.hide()
            self.ruta_factura.clear() 

    def seleccionar_factura(self):
        ruta, _ = QFileDialog.getOpenFileName(self, 'Seleccionar Factura', '', 'Archivos (*.pdf)')
        if ruta:
            self.ruta_factura.setText(ruta)
    def toggle_oneroso(self, state):
        if state == Qt.Checked:
            self.combo_oneroso.show()
            self.cargar_opciones_oneroso()
        else:
            self.combo_oneroso.hide()
            self.combo_oneroso.clear()
            self.combo_oneroso.addItem('Seleccione una opción')
    def validar_correo(self):
        texto = self.campo_correo.text()
        patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if re.match(patron, texto):
            self.campo_correo.setStyleSheet("border: 2px solid green;")
        else:
            self.campo_correo.setStyleSheet("border: 2px solid red;")

    def seleccionar_archivo_cedula(self):
        ruta, _ = QFileDialog.getOpenFileName(self, 'Seleccionar Cédula', '', 'Archivos (*.pdf *.png *.jpg *.jpeg)')
        if ruta:
            self.ruta_cedula.setText(ruta)

    def seleccionar_archivo_tarjeta(self):
        ruta, _ = QFileDialog.getOpenFileName(self, 'Seleccionar Tarjeta de Propiedad', '', 
                                              'Archivos (*.pdf *.png *.jpg *.jpeg)')
        if ruta:
            self.ruta_tarjeta.setText(ruta)

    def actualizar_ciudades(self):
        self.combo_ciudad.clear()
        self.combo_ciudad.addItem('Seleccione una ciudad')

        index = self.combo_departamento.currentIndex()
        if index > 0:  
            departamento = self.departamentos[index - 1]
            ciudades = departamento['ciudades']
            self.label_departamento.setText(f"Zona de circulación ({departamento['departamento']})")

            for ciudad in ciudades:
                self.combo_ciudad.addItem(ciudad)
        else:
            self.label_departamento.setText('Zona de circulación:')

    def guardar_datos_en_excel(self, datos):
        ruta_datos_excel = 'datos/datos_extraidos.xlsx'
        encabezados = ['NUM_CEDULA','NOMBRES', 'APELLIDOS', 'FECHA_NACIMIENTO', 'FECHA_LUGAR_EXPEDICION', 'CORREO',
                       'CLASE_VEHICULO', 'PLACA', 'MODELO', 'SERVICIO', 'MARCA', 'LÍNEA', 'CILINDRAJE', 'VEHICULO_NUEVO(0KM)', 'DEP_CIRCULACION', 
                       'CIUDAD_CIRCULACIÓN', 'GÉNERO', 'B.ONEROSO']

        if os.path.exists(ruta_datos_excel):
            wb = load_workbook(ruta_datos_excel)
            ws = wb.active
            ws.title = "Datos Extraídos"

        else:
            wb = Workbook()
            ws = wb.active

            for col_num, encabezado in enumerate(encabezados, start=1):
                celda = ws.cell(row=1, column=col_num, value=encabezado)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25

        cedula = datos.get('cedula', {})
        tarjeta = datos.get('tarjeta', {})

        num_cedula = cedula.get('num_cedula', 'N/A')
        nombres = cedula.get('nombres', 'N/A')
        apellidos = cedula.get('apellidos', 'N/A')
        fecha_nacimiento = cedula.get('fecha_nacimiento', 'N/A')
        fecha_lugar_expedicion = cedula.get('fecha_lugar_expedicion', 'N/A')

        placa = tarjeta.get('placa', 'N/A')
        clase_vehiculo = tarjeta.get('clase_vehiculo', 'N/A')
        modelo = tarjeta.get('modelo', 'N/A')
        servicio = tarjeta.get('servicio', 'N/A')
        marca = tarjeta.get('marca', 'N/A')
        linea = tarjeta.get('linea', 'N/A')
        cilindraje = tarjeta.get('cilindraje', 'N/A')

        nuevo_vehiculo = datos.get('nuevo_vehiculo', 'N/A')
        zona = datos.get('zona_circulacion', {})
        zona_departamento = zona.get('departamento', 'N/A')
        zona_ciudad = zona.get('ciudad', 'N/A')
        genero = datos.get('genero', 'N/A')
        correo = datos.get('correo', 'N/A')
        oneroso = datos.get('oneroso', 'N/A')

        ws.append([num_cedula, nombres, apellidos, fecha_nacimiento, fecha_lugar_expedicion, correo, clase_vehiculo, 
                   placa, modelo, servicio, marca, linea, cilindraje, nuevo_vehiculo, zona_departamento, zona_ciudad, genero, oneroso])

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(ruta_datos_excel)
        print("DATOS GUARDADOS EN EXCEL CORRECTAMENTE")

    def mostrar_alerta(self, titulo, mensaje, tipo="Warning"):
        alerta = QMessageBox()

        if tipo == "Information":
            alerta.setIcon(QMessageBox.Information)
        elif tipo == "Critical":
            alerta.setIcon(QMessageBox.Critical)
        else:
            alerta.setIcon(QMessageBox.Warning) 

        alerta.setWindowTitle(titulo)
        alerta.setText(mensaje)
        alerta.setStandardButtons(QMessageBox.Ok)
        alerta.exec_()

    def procesar_formulario(self):
        cedula = self.ruta_cedula.text()
        tarjeta = self.ruta_tarjeta.text()
        departamento = self.combo_departamento.currentText()
        ciudad = self.combo_ciudad.currentText()
        genero = self.combo_genero.currentText()
        correo = self.campo_correo.text()

        texto = self.campo_correo.text()
        patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(patron, texto):
            QMessageBox.warning(self, "Error", "Por favor, ingrese un correo válido.")
            return

        if self.checkbox_oneroso.isChecked() and self.combo_oneroso.currentIndex() == 0:
            self.mostrar_alerta('Error', 'Debes seleccionar una opción en el campo de Beneficiario Oneroso.', "Warning")
            return

        oneroso = self.combo_oneroso.currentText() if self.checkbox_oneroso.isChecked() else "NO"
        nuevo_vehiculo = "NO"

        if self.checkbox_nuevo.isChecked():
            nuevo_vehiculo = self.ruta_factura.text()
            if not nuevo_vehiculo:
                self.mostrar_alerta('Error', 'Por favor, suba el archivo de la factura del vehículo', "Warning")
                return

        if not cedula:
            self.mostrar_alerta('Error', 'Por favor, adjunte el archivo de cédula.', "Warning")
            return

        if not tarjeta:
            self.mostrar_alerta('Error', 'Por favor, adjunte el archivo de tarjeta de propiedad.', "Warning")
            return

        if not departamento or departamento == 'Seleccione un departamento':
            self.mostrar_alerta('Error', 'Por favor, seleccione un departamento.', "Warning")
            return

        if not ciudad or ciudad == 'Seleccione una ciudad':
            self.mostrar_alerta('Error', 'Por favor, seleccione una ciudad.', "Warning")  
            return

        if not genero or genero == 'Seleccione el genero del propietario':
            self.mostrar_alerta('Error', 'Por favor, seleccione el genero del propietario.', "Warning")
            return

        if not correo or correo == 'Escribe un correo electrónico':
            self.mostrar_alerta('Error', 'Por favor, escriba un correo electrónico.', "Warning")
            return

        self.label_estado.setText("Procesando archivos, por favor espere...")
        self.label_estado.show()
        QApplication.processEvents()
        try:
            datos_extraidos = procesar_documentos(cedula, tarjeta)
            datos_extraidos['nuevo_vehiculo'] = nuevo_vehiculo
            datos_extraidos['zona_circulacion'] = {
                'departamento': departamento,
                'ciudad': ciudad
            }
            datos_extraidos['genero'] = genero
            datos_extraidos['correo'] = correo
            datos_extraidos['oneroso'] = oneroso

            if datos_extraidos:
                texto_tarjeta = datos_extraidos.get("texto_tarjeta", "No se pudo extraer texto")

                ruta_datos = 'datos/datos_extraidos.json'
                os.makedirs(os.path.dirname(ruta_datos), exist_ok=True)
                with open(ruta_datos, 'w', encoding='utf-8') as archivo_json:
                    json.dump(datos_extraidos, archivo_json, ensure_ascii=False, indent=4)

                self.guardar_datos_en_excel(datos_extraidos)
                self.mostrar_alerta('Éxito', 'Los datos se han procesado y guardado correctamente.', "Information")

                self.ruta_cedula.clear()
                self.ruta_tarjeta.clear()
                self.combo_departamento.setCurrentIndex(0)
                self.ruta_factura.clear()
                self.checkbox_nuevo.setChecked(False)
                self.combo_ciudad.clear()
                self.combo_ciudad.addItem('Seleccione una ciudad')
                self.combo_genero.setCurrentIndex(0)
                self.campo_correo.clear()
                self.combo_oneroso.setCurrentIndex(0)
                self.checkbox_oneroso.setChecked(False)

                try:
                    subprocess.Popen(['python', 'web.py'], shell=True)
                    print("Iniciando automatización en Agente Motor")
                except Exception as e:
                    self.mostrar_alerta('Error', f'No se puedo iniciar automatización en Agente Motor: {e}', "Critical")
        except Exception as e:
            self.mostrar_alerta('Error', f'Ocurrió un error al procesar los archivos: {e}', "Critical")
        finally:
            self.label_estado.hide()

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    formulario = FormularioApp()
    formulario.show()
    sys.exit(app.exec_())